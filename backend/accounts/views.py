import json
from io import BytesIO

from django.contrib.auth import get_user_model, login, logout
from django.contrib.auth.models import Group
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook


def _user_payload(user):
	primary_group = user.groups.first()
	return {
		"id": user.id,
		"username": user.get_username(),
		"email": user.email,
		"role": primary_group.name if primary_group else None,
	}

# Login Function
@csrf_exempt
@require_POST
def login_view(request):
	try:
		payload = json.loads(request.body or "{}")
	except json.JSONDecodeError:
		return JsonResponse({"detail": "Invalid JSON payload."}, status=400)

	email = (payload.get("email") or "").strip().lower()
	password = payload.get("password") or ""
	remember_me = bool(payload.get("rememberMe"))

	if not email or not password:
		return JsonResponse({"detail": "Email and password are required."}, status=400)

	UserModel = get_user_model()
	users = list(UserModel.objects.filter(email__iexact=email)[:2])

	if not users:
		return JsonResponse({"detail": "Invalid credentials."}, status=401)

	if len(users) > 1:
		return JsonResponse(
			{"detail": "Multiple accounts share this email. Contact support."},
			status=409,
		)

	user = users[0]

	if not user.is_active:
		return JsonResponse({"detail": "Account is inactive."}, status=403)

	if not user.check_password(password):
		return JsonResponse({"detail": "Invalid credentials."}, status=401)

	user.backend = "django.contrib.auth.backends.ModelBackend"
	login(request, user)
	if remember_me:
		request.session.set_expiry(60 * 60 * 24 * 30)  # 30 days
	else:
		request.session.set_expiry(0)  # Browser session

	return JsonResponse({"user": _user_payload(user)})

# User List display Function
@csrf_exempt
@require_GET
def user_list_view(request):
	if not request.user.is_authenticated:
		return JsonResponse({"detail": "Not authenticated."}, status=401)
	
	if not request.user.is_authenticated or not request.user.groups.filter(name='admin').exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	UserModel = get_user_model()
	users = UserModel.objects.all()
	user_list = [_user_payload(user) for user in users]

	return JsonResponse({"users": user_list})

# Logout Function
@csrf_exempt
@require_POST
def logout_view(request):
	logout(request)
	return JsonResponse({"success": True})

# Retrieve Current User information Function
@require_GET
def current_user_view(request):
	if not request.user.is_authenticated:
		return JsonResponse({"detail": "Not authenticated."}, status=401)

	return JsonResponse({"user": _user_payload(request.user)})

# Create New User Function
@csrf_exempt
@require_POST
def create_user_view(request):
	if not request.user.is_authenticated or not request.user.groups.filter(name='admin').exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	try:
		payload = json.loads(request.body or "{}")
	except json.JSONDecodeError:
		return JsonResponse({"detail": "Invalid JSON payload."}, status=400)

	username = (payload.get("username") or "").strip()
	email = (payload.get("email") or "").strip().lower()
	password = payload.get("password") or ""
	role = (payload.get("role") or "").strip().lower()

	if not username or not email or not password or role not in {"admin", "editor", "viewer"}:
		return JsonResponse({"detail": "Username, email, password, and valid role are required."}, status=400)

	UserModel = get_user_model()
	if UserModel.objects.filter(email__iexact=email).exists():
		return JsonResponse({"detail": "Email already in use."}, status=409)

	user = UserModel.objects.create_user(username=username, email=email, password=password)
	group = Group.objects.get(name=role)
	user.groups.add(group)
	user.save()

	return JsonResponse({"user": _user_payload(user)}, status=201)


# Update Existing User Function
@csrf_exempt
@require_POST
def update_user_view(request, user_id):
	if not request.user.is_authenticated or not request.user.groups.filter(name='admin').exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	try:
		payload = json.loads(request.body or "{}")
	except json.JSONDecodeError:
		return JsonResponse({"detail": "Invalid JSON payload."}, status=400)

	username = (payload.get("username") or "").strip()
	email = (payload.get("email") or "").strip().lower()
	password = payload.get("password")
	role = (payload.get("role") or "").strip().lower()

	if not username or not email or role not in {"admin", "editor", "viewer"}:
		return JsonResponse({"detail": "Username, email, and valid role are required."}, status=400)

	UserModel = get_user_model()
	try:
		user = UserModel.objects.get(id=user_id)
	except UserModel.DoesNotExist:
		return JsonResponse({"detail": "User not found."}, status=404)

	if UserModel.objects.filter(email__iexact=email).exclude(id=user_id).exists():
		return JsonResponse({"detail": "Email already in use."}, status=409)

	user.username = username
	user.email = email
	if password:
		user.set_password(password)
	group = Group.objects.get(name=role)
	user.groups.clear()
	user.groups.add(group)
	user.save()

	return JsonResponse({"user": _user_payload(user)})

# Delete User Function
@csrf_exempt
@require_POST
def delete_user_view(request, user_id):
	if not request.user.is_authenticated or not request.user.groups.filter(name='admin').exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	UserModel = get_user_model()
	try:
		user = UserModel.objects.get(id=user_id)
	except UserModel.DoesNotExist:
		return JsonResponse({"detail": "User not found."}, status=404)

	user.delete()
	return JsonResponse({"success": True})


@csrf_exempt
@require_GET
def user_import_sample_view(request):
	if not request.user.is_authenticated or not request.user.groups.filter(name='admin').exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	wb = Workbook()
	ws = wb.active
	ws.title = "Users"
	ws.append(["username", "email", "password", "role"])
	ws.append(["testing", "testing@example.com", "123", "viewer"])

	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)

	response = HttpResponse(
		buffer.getvalue(),
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="User_import_sample.xlsx"'
	return response


@csrf_exempt
@require_POST
def bulk_import_users_view(request):
	if not request.user.is_authenticated or not request.user.groups.filter(name="admin").exists():
		return JsonResponse({"detail": "Not authorized."}, status=403)

	upload = request.FILES.get("file")
	if not upload:
		return JsonResponse({"detail": "File required."}, status=400)

	wb = load_workbook(upload, data_only=True)
	ws = wb.active

	required = ["username", "email", "password", "role"]
	header_row = next(ws.rows, None)
	headers = [str(cell.value or "").strip().lower() for cell in header_row] if header_row else []
	if headers != required:
		return JsonResponse({"detail": f"Expected header row {required}."}, status=400)

	User = get_user_model()
	role_groups = {group.name.lower(): group for group in Group.objects.filter(name__in=["admin", "editor", "viewer"])}
	created, skipped_existing, skipped_invalid = [], [], []
	seen = set()
	with transaction.atomic():
		for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			username_raw, email_raw, password_raw, role_raw = (str(value or "").strip() for value in (row or (None, None, None, None)))

			if not any([username_raw, email_raw, password_raw, role_raw]):
				# Ignore completely empty rows
				continue

			email_lower = email_raw.lower()
			role_lower = role_raw.lower()

			if not username_raw or not email_lower or not password_raw or role_lower not in role_groups:
				skipped_invalid.append({"row": idx, "reason": "Missing fields or invalid role"})
				continue
			if email_lower in seen:
				skipped_invalid.append({"row": idx, "reason": "Duplicate email in file"})
				continue
			if User.objects.filter(email__iexact=email_lower).exists():
				skipped_existing.append({"row": idx, "email": email_lower})
				continue

			user = User.objects.create_user(username=username_raw, email=email_lower, password=password_raw)
			user.groups.add(role_groups[role_lower])
			created.append(user.id)
			seen.add(email_lower)

	return JsonResponse({
		"createdCount": len(created),
		"skippedExisting": skipped_existing,
		"skippedInvalid": skipped_invalid,
	}, status=201)